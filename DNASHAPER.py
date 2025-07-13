import streamlit as st
import zipfile
import pandas as pd
from io import StringIO, BytesIO
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl import Workbook
import csv

st.title("üß¨ DNA ShapeR Excel Generator")

uploaded_zip = st.file_uploader("Upload ZIP file containing `.txt` files (from DNAShapeR)", type=["zip"])
uploaded_fasta = st.file_uploader("Upload the corresponding `.fasta` file", type=["fasta", "fa", "txt"])

if uploaded_zip and uploaded_fasta:
    st.success("‚úÖ Both ZIP and FASTA uploaded!")

    # --- Parse FASTA (multiline-safe) ---
    fasta_lines = uploaded_fasta.read().decode('utf-8').splitlines()
    sequence_ids, sequences = [], []
    current_id, current_seq_lines = None, []

    for line in fasta_lines:
        if line.startswith(">"):
            if current_id is not None:
                sequence_ids.append(current_id)
                sequences.append(''.join(current_seq_lines))
            current_id = line[1:].strip()
            current_seq_lines = []
        else:
            current_seq_lines.append(line.strip())
    if current_id is not None:
        sequence_ids.append(current_id)
        sequences.append(''.join(current_seq_lines))

    # --- Read and clean TXT files ---
    zip_bytes = BytesIO(uploaded_zip.read())
    dataframes = {}
    row_counts = []

    with zipfile.ZipFile(zip_bytes, 'r') as zip_ref:
        txt_files = [f for f in zip_ref.namelist() if f.endswith('.txt')]

        for file_name in txt_files:
            with zip_ref.open(file_name) as file:
                lines = [line.decode('utf-8').strip().replace('\t', ' ') for line in file.readlines()]
                cleaned_lines = [' '.join(line.split()) for line in lines]
                csv_ready = '\n'.join([line.replace(' ', ',') for line in cleaned_lines])

                reader = csv.reader(StringIO(csv_ready))
                rows = list(reader)
                rows = [r for r in rows if len(r) > 0 and not all(cell == '' for cell in r)]
                max_cols = max(len(row) for row in rows)
                padded_rows = [row + ['0'] * (max_cols - len(row)) for row in rows]
                if any(any(c.isalpha() for c in cell) for cell in padded_rows[0]):
                    padded_rows = padded_rows[1:]  # skip header row
                padded_rows = [row[1:] for row in padded_rows]  # remove first column (srno)
                df = pd.DataFrame(padded_rows)
                df = df.apply(pd.to_numeric, errors='coerce').fillna(0)

                base_name = file_name.split('/')[-1].replace('.txt', '')
                df[f"avg({base_name})"] = df.mean(axis=1)
                dataframes[base_name] = df
                row_counts.append(len(df))

    # --- Auto-trim if 1 row mismatch
    min_len = min(row_counts[0], len(sequence_ids))
    if abs(len(sequence_ids) - row_counts[0]) == 1:
        st.warning("‚ö†Ô∏è FASTA and TXT row mismatch by 1 ‚Äî auto-trimmed.")
        sequence_ids = sequence_ids[:min_len]
        sequences = sequences[:min_len]
        for key in dataframes:
            dataframes[key] = dataframes[key].iloc[:min_len]
        row_counts = [min_len]

    # --- Validation
    if len(set(row_counts)) != 1:
        st.error("‚ùå Not all `.txt` files have the same number of rows.")
        for fname, df in dataframes.items():
            st.text(f"{fname}: {df.shape[0]} rows")
        st.stop()
    if len(sequence_ids) != row_counts[0]:
        st.error(f"‚ùå FASTA sequence count ({len(sequence_ids)}) does not match data rows ({row_counts[0]}).")
        st.stop()

    # Styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    blue_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    avg_font = Font(color="0070C0", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Create Excel workbook
    output = BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Combined Data"

    # Header rows
    ws1.cell(row=2, column=1, value="Sequence ID").fill = blue_fill
    ws1.cell(row=2, column=2, value="Sequence").fill = blue_fill
    ws1.cell(row=2, column=1).alignment = center_align
    ws1.cell(row=2, column=2).alignment = center_align
    ws1.cell(row=2, column=1).font = bold_font
    ws1.cell(row=2, column=2).font = bold_font

    start_col = 3
    for df_name, df in dataframes.items():
        col_count = df.shape[1]
        col_start = start_col
        col_end = col_start + col_count - 1

        ws1.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        header_cell = ws1.cell(row=1, column=col_start, value=df_name)
        header_cell.fill = header_fill
        header_cell.alignment = center_align
        header_cell.font = bold_font

        for j in range(df.shape[1] - 1):
            sub_col = ws1.cell(row=2, column=col_start + j, value=f"{df_name}_{j+1}")
            sub_col.fill = blue_fill
            sub_col.alignment = center_align
            sub_col.font = bold_font
            sub_col.border = thin_border

        avg_col = ws1.cell(row=2, column=col_end, value=f"avg({df_name})")
        avg_col.fill = blue_fill
        avg_col.font = avg_font
        avg_col.alignment = center_align
        avg_col.border = thin_border

        start_col += col_count

    # Fill data
    for i in range(row_counts[0]):
        ws1.cell(row=i + 3, column=1, value=sequence_ids[i]).font = bold_font
        ws1.cell(row=i + 3, column=2, value=sequences[i]).font = bold_font
        ws1.cell(row=i + 3, column=1).border = thin_border
        ws1.cell(row=i + 3, column=2).border = thin_border

    start_col = 3
    for df in dataframes.values():
        for i, row in df.iterrows():
            for j, val in enumerate(row):
                cell = ws1.cell(row=i + 3, column=start_col + j, value=val)
                cell.border = thin_border
                if j == len(row) - 1:
                    cell.font = avg_font
        start_col += df.shape[1]
    # Create new worksheet for only avg values
    ws_avg = wb.create_sheet("Only Averages")


    # Header
    ws_avg.cell(row=1, column=1, value="Sequence ID").fill = blue_fill
    ws_avg.cell(row=1, column=2, value="Sequence").fill = blue_fill
    ws_avg.cell(row=1, column=1).alignment = center_align
    ws_avg.cell(row=1, column=2).alignment = center_align
    ws_avg.cell(row=1, column=1).font = bold_font
    ws_avg.cell(row=1, column=2).font = bold_font

    # Add avg column headers
    avg_headers = list(dataframes.keys())
    for col_idx, df_name in enumerate(avg_headers, start=3):
        cell = ws_avg.cell(row=1, column=col_idx, value=f"avg({df_name})")
        cell.fill = blue_fill
        cell.font = avg_font
        cell.alignment = center_align
        cell.border = thin_border

    # Fill data rows
    for i in range(row_counts[0]):
        ws_avg.cell(row=i + 2, column=1, value=sequence_ids[i]).font = bold_font
        ws_avg.cell(row=i + 2, column=2, value=sequences[i]).font = bold_font
        ws_avg.cell(row=i + 2, column=1).border = thin_border
        ws_avg.cell(row=i + 2, column=2).border = thin_border

        for col_idx, df in enumerate(dataframes.values(), start=3):
            avg_val = df.iloc[i, -1]
            cell = ws_avg.cell(row=i + 2, column=col_idx, value=avg_val)
            cell.font = avg_font
            cell.border = thin_border

    wb.save(output)
    output.seek(0)

    st.success("‚úÖ Styled Excel with colors, borders, and headers generated.")
    st.download_button("üìÖ Download Styled Excel", output, file_name="DNAshape_colored.xlsx")